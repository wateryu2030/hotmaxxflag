# -*- coding: utf-8 -*-
"""
开票台账三类表格：每日未开票收入计算表、每日开票台账、每日收入汇总。
生成空白模板（含表头与示例备注行），供 Excel / PDF 下载。
"""
from __future__ import annotations

import io
from typing import List

DEFAULT_STORE_NAME = "好特卖超级仓沈阳HAI乐园店"


def build_invoicing_ledger_xlsx(store_name: str | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    store = (store_name or DEFAULT_STORE_NAME).strip() or DEFAULT_STORE_NAME
    wb = Workbook()
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ----- Sheet1 -----
    ws1 = wb.active
    ws1.title = "未开票收入计算"
    headers1 = [
        "月份",
        "本月总营收（不含税）",
        "本月总开票金额（不含税）",
        "本月未开票收入（不含税）",
        "本月补开前期发票金额（不含税）",
        "申报表未开票栏填写金额",
        "备注（补开对应月份，异常说明等）",
    ]
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers1))
    c = ws1.cell(1, 1, "每日未开票收入计算表")
    c.font = title_font
    c.alignment = center
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers1))
    ws1.cell(2, 1, f"卖场：{store}").alignment = left
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(3, col, h)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for r in range(4, 20):
        for col in range(1, len(headers1) + 1):
            ws1.cell(r, col, "").border = border
    ws1.cell(4, 7, "无补开，全部合规").alignment = left
    ws1.cell(5, 7, "补开2026年1月未开票收入").alignment = left
    widths1 = [10, 18, 20, 22, 26, 22, 40]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ----- Sheet2 -----
    ws2 = wb.create_sheet("每日开票台账")
    headers2 = [
        "序号",
        "开票日期",
        "发票号码",
        "购买方名称",
        "税号",
        "价税合计",
        "不含税金额",
        "税额",
        "开票人",
        "开票归属月份",
        "流水交易码",
        "备注1：不含税",
        "备注2：税额13%",
        "备注3：不含税",
        "备注4：税额9%",
        "备注5：不含税",
        "备注6：零税率",
    ]
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers2))
    ws2.cell(1, 1, "每日开票台账").font = title_font
    ws2.cell(1, 1).alignment = center
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers2))
    ws2.cell(2, 1, f"卖场：{store}").alignment = left
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(3, col, h)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for r in range(4, 28):
        for col in range(1, len(headers2) + 1):
            ws2.cell(r, col, "").border = border
    for i in range(1, 18):
        ws2.column_dimensions[get_column_letter(i)].width = 11 if i > 11 else 13

    # ----- Sheet3 -----
    ws3 = wb.create_sheet("每日收入汇总")
    ncols = 14
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws3.cell(1, 1, "每日收入汇总").font = title_font
    ws3.cell(1, 1).alignment = center
    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws3.cell(2, 1, f"卖场：{store}").alignment = left

    ws3.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    ws3.cell(3, 1, "序号").alignment = center
    ws3.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    ws3.cell(3, 2, "日期").alignment = center
    ws3.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)
    ws3.cell(3, 3, "小计(1+2+3)").alignment = center
    ws3.merge_cells(start_row=3, start_column=6, end_row=3, end_column=8)
    ws3.cell(3, 6, "13%(1)").alignment = center
    ws3.merge_cells(start_row=3, start_column=9, end_row=3, end_column=11)
    ws3.cell(3, 9, "9%(2)").alignment = center
    ws3.merge_cells(start_row=3, start_column=12, end_row=3, end_column=14)
    ws3.cell(3, 12, "零税率(3)").alignment = center
    sub = ["含税", "不含税", "税额"]
    for base in (3, 6, 9, 12):
        for i, lab in enumerate(sub):
            ws3.cell(4, base + i, lab).alignment = center
    for r in (3, 4):
        for col in range(1, ncols + 1):
            ws3.cell(r, col).font = header_font
            ws3.cell(r, col).border = border
    for r in range(5, 22):
        for col in range(1, ncols + 1):
            ws3.cell(r, col, "").border = border
    ws3.cell(5, 1, 1)
    ws3.cell(5, 2, "12月25日")
    for col in range(1, ncols + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_invoicing_ledger_pdf(store_name: str | None = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    store = (store_name or DEFAULT_STORE_NAME).strip() or DEFAULT_STORE_NAME
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    story: List = []
    styles = getSampleStyleSheet()
    fn = "Helvetica"
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont

        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        fn = "STSong-Light"
    except Exception:
        pass

    h1 = styles["Heading1"]
    h1.fontName = fn
    nb = styles["Normal"]
    nb.fontName = fn

    def add_title(title: str):
        story.append(Paragraph(title.replace("&", "&amp;"), h1))
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(f"卖场：{store.replace('&', '&amp;')}", nb))
        story.append(Spacer(1, 4 * mm))

    def add_table(data: List[List[str]], ncols: int, fs: int = 7):
        cw = doc.width / ncols
        t = Table(data, colWidths=[cw] * ncols)
        t.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), fs),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, -1), fn),
                ]
            )
        )
        story.append(t)
        story.append(Spacer(1, 10 * mm))

    # 表一
    add_title("每日未开票收入计算表")
    h1_row = [
        "月份",
        "本月总营收(不含税)",
        "本月总开票(不含税)",
        "本月未开票收入(不含税)",
        "本月补开前期(不含税)",
        "申报未开票栏",
        "备注",
    ]
    d1 = [h1_row]
    d1.append(["", "", "", "", "", "", "无补开，全部合规"])
    d1.append(["", "", "", "", "", "", "补开2026年1月未开票收入"])
    d1.extend([[""] * 7 for _ in range(12)])
    add_table(d1, 7, 8)

    # 表二（新页由分页自然产生）
    story.append(Spacer(1, 4 * mm))
    add_title("每日开票台账")
    h2 = [
        "序号",
        "开票日期",
        "发票号码",
        "购买方名称",
        "税号",
        "价税合计",
        "不含税金额",
        "税额",
        "开票人",
        "开票归属月份",
        "流水交易码",
        "备注1不含税",
        "备注213%税",
        "备注3不含税",
        "备注49%税",
        "备注5不含税",
        "备注6零税率",
    ]
    d2 = [h2] + [[""] * 17 for _ in range(20)]
    add_table(d2, 17, 6)

    add_title("每日收入汇总")
    h3 = [
        "序号",
        "日期",
        "小计含税",
        "小计不含税",
        "小计税额",
        "13%含税",
        "13%不含税",
        "13%税额",
        "9%含税",
        "9%不含税",
        "9%税额",
        "零税率含税",
        "零税率不含税",
        "零税率税额",
    ]
    d3 = [h3, ["1", "12月25日"] + [""] * 12]
    d3.extend([["", ""] + [""] * 12 for _ in range(14)])
    add_table(d3, 14, 7)

    doc.build(story)
    return buf.getvalue()
